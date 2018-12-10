#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import load_workbook

# 引入xlsx文件
wb = load_workbook("F:\\python\\示例\\操作excel\\123.xlsx")

# 打印当前表
# print(wb.active)

ws=wb.active
ws.title = "New"

# 创建新表默认插在工作簿末尾
ws1 = wb.create_sheet()

# 打印所有的sheet名字
# print(wb.sheetnames) 

# 获取某个单元格
b4 = wb.active['B4']
print(b4)
print(f'({b4.column},{b4.row}) is {b4.value}')

# 第二种方式获取某个单元格
b4_too = wb.active.cell(row=4 ,column=2)
print(b4_too.value)

# 获得最大列和最大行
print(wb.active.max_row)
print(wb.active.max_column)










